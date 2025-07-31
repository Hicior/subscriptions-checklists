import os
import requests
import pandas as pd
import datetime
import time
from datetime import datetime as dt, timezone
from tqdm import tqdm
from openpyxl import load_workbook
import numpy as np

# =============================================================================
# CONFIGURATION AND API SETUP
# =============================================================================

# API Keys
calendesk_api_key = os.getenv('CALENDESK_API_KEY')
stripe_api_key = os.getenv('STRIPE_API_KEY')

if not calendesk_api_key:
    print("❌ No Calendesk API key found. Please set the CALENDESK_API_KEY environment variable.")
    exit()

if not stripe_api_key:
    print("❌ No Stripe API key found. Please set the STRIPE_API_KEY environment variable.")
    exit()

print("✅ API keys loaded successfully")

# Excel file configuration
excel_file_path = os.path.join('ExcelFiles', 'Baza subskrypji - checklisty.xlsx')
calendesk_sheet_name = 'CalendeskSubs'
stripe_sheet_name = 'StripeInvoices'

# Calendesk API configuration
calendesk_headers = {
    "X-Tenant": "slawomir-mentzen-rvs",
    "X-Api-Key": calendesk_api_key
}

subscriptions_url = 'https://api.calendesk.com/api/admin/subscriptions'
users_url = 'https://api.calendesk.com/api/admin/v2/users/subscriptions'

# Stripe API configuration
stripe_headers = {
    'Authorization': f'Bearer {stripe_api_key}',
}
stripe_base_url = 'https://api.stripe.com/v1/invoices'

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def fetch_calendesk_data(url, headers, description="Fetching data"):
    """Fetch all data from a Calendesk API endpoint using dynamic pagination"""
    start_time = time.time()
    all_data = []
    current_page = 1
    total_pages = None
    total_records = None
    
    print(f"📡 Starting to fetch data from {url}")
    
    # First request to get pagination info
    params = {
        'limit': 100,  # Hardcoded to 100 as per requirement
        'page': current_page,
        'order_by': 'id',
        'ascending': 0
    }
    
    response = make_api_request_with_retry(url, headers, params)
    if not response or response.status_code != 200:
        error_msg = f"Status code {response.status_code}" if response else "No response after retries"
        print(f"❌ Failed to fetch initial data: {error_msg}")
        return []
    
    response_data = response.json()
    total_pages = response_data.get('last_page', 1)
    total_records = response_data.get('total', 0)
    per_page = response_data.get('per_page', 50)
    current_page_data = response_data.get('data', [])
    
    # Add first page data
    all_data.extend(current_page_data)
    
    print(f"📊 Found {total_records} total records across {total_pages} pages")
    print(f"📈 API limit: 100, actual per_page: {per_page}, first page: {len(current_page_data)} records")
    
    # If we only have one page, return early
    if total_pages <= 1:
        print(f"✓ Single page fetch completed: {len(all_data)} records")
        return all_data
    
    # Continue with remaining pages
    current_page = 2
    
    with tqdm(total=total_pages, initial=1, desc=description, unit="page") as pbar:
        while current_page <= total_pages:
            params = {
                'limit': 100,  # Hardcoded to 100 as per requirement
                'page': current_page,
                'order_by': 'id',
                'ascending': 0
            }
            
            response = make_api_request_with_retry(url, headers, params)
            if response and response.status_code == 200:
                response_data = response.json()
                data = response_data.get('data', [])
                
                if data:
                    all_data.extend(data)
                    pbar.update(1)
                    
                    # Show progress every 10 pages or for smaller datasets every page
                    if current_page % 10 == 0 or total_pages <= 20:
                        print(f"  ✓ Page {current_page}/{total_pages}: {len(data)} records ({len(all_data)} total)")
                    
                    # Update pagination info (in case it changes during fetching)
                    updated_total_pages = response_data.get('last_page', total_pages)
                    if updated_total_pages != total_pages:
                        print(f"  📊 Updated total pages: {updated_total_pages} (was {total_pages})")
                        total_pages = updated_total_pages
                        pbar.total = total_pages
                    
                    # Check if there's a next page
                    next_page_url = response_data.get('next_page_url')
                    if not next_page_url and current_page >= total_pages:
                        print(f"  ✓ Reached end of data at page {current_page}")
                        break
                        
                    current_page += 1
                else:
                    print(f"  ⚠ Page {current_page}: No data returned, stopping")
                    break
            else:
                error_msg = f"Status code {response.status_code}" if response else "No response after retries"
                print(f'  ❌ Failed to fetch page {current_page}: {error_msg}')
                if response:
                    print(f"     Response: {response.text[:200]}...")
                break
    
    print(f"✓ Total records fetched: {len(all_data)} out of {total_records} expected")
    
    # Verify we got all expected records
    if len(all_data) != total_records:
        missing = total_records - len(all_data)
        print(f"⚠ Warning: Expected {total_records} records but got {len(all_data)} (missing {missing})")
        
        # If we're significantly short, warn about potential data loss
        if missing > len(all_data) * 0.1:  # More than 10% missing
            print(f"🚨 Significant data missing - consider re-running the fetch")
    else:
        print(f"✅ Successfully fetched all expected records!")
    
    # Final validation on complete dataset
    if len(all_data) > 0:
        print(f"📊 Dataset summary:")
        print(f"   - Total records: {len(all_data)}")
        print(f"   - Sample fields: {list(all_data[0].keys())[:8]}...")
        
        # Check for duplicate IDs
        if 'id' in all_data[0]:
            ids = [record['id'] for record in all_data]
            unique_ids = set(ids)
            if len(ids) != len(unique_ids):
                duplicates = len(ids) - len(unique_ids)
                print(f"⚠ Warning: Found {duplicates} duplicate IDs in dataset")
    
    # Performance metrics
    end_time = time.time()
    duration = end_time - start_time
    records_per_second = len(all_data) / duration if duration > 0 else 0
    print(f"⏱ Fetch completed in {duration:.1f} seconds ({records_per_second:.1f} records/sec)")
    
    return all_data

def convert_timestamp_to_date(timestamp):
    """Convert Unix timestamp to date string"""
    return dt.fromtimestamp(timestamp).strftime('%Y-%m-%d')

def make_api_request_with_retry(url, headers, params, max_retries=3, delay=2):
    """Make API request with retry logic for transient failures"""
    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                return response
            elif response.status_code == 429:  # Rate limiting
                wait_time = delay * (2 ** attempt)  # Exponential backoff
                print(f"  ⏳ Rate limited (attempt {attempt + 1}/{max_retries}), waiting {wait_time} seconds...")
                time.sleep(wait_time)
                continue
            elif response.status_code >= 500:  # Server errors
                wait_time = delay * (attempt + 1)
                print(f"  🔄 Server error {response.status_code} (attempt {attempt + 1}/{max_retries}), retrying in {wait_time} seconds...")
                time.sleep(wait_time)
                continue
            else:
                # Client error, don't retry
                return response
                
        except requests.exceptions.Timeout:
            print(f"  ⏱ Request timeout (attempt {attempt + 1}/{max_retries})")
            if attempt < max_retries - 1:
                time.sleep(delay * (attempt + 1))
                continue
        except requests.exceptions.RequestException as e:
            print(f"  ❌ Request error (attempt {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(delay * (attempt + 1))
                continue
    
    # All retries failed
    return None

def validate_calendesk_data(data, endpoint_type):
    """Validate the structure of Calendesk API data"""
    if not data:
        print(f"⚠ Warning: No data returned from {endpoint_type} endpoint")
        return False
    
    required_fields = {
        'subscriptions': ['id', 'name', 'price'],
        'users': ['id', 'subscription_id', 'user', 'status']
    }
    
    fields_to_check = required_fields.get(endpoint_type, [])
    if not fields_to_check:
        return True
    
    # Check first few records for required fields
    sample_size = min(3, len(data))
    missing_fields = set()
    
    for i in range(sample_size):
        record = data[i]
        for field in fields_to_check:
            if field not in record:
                missing_fields.add(field)
    
    if missing_fields:
        print(f"⚠ Warning: Missing expected fields in {endpoint_type} data: {missing_fields}")
        print(f"   Sample record keys: {list(data[0].keys())[:10]}...")
        return False
    
    print(f"✅ Data structure validation passed for {endpoint_type}")
    return True

def fetch_stripe_invoices():
    """Fetch all Stripe invoices with pagination"""
    # Calculate Unix timestamp for June 1st, 2024 (to get more historical data)
    filter_date = dt(2025, 6, 1, 0, 0, 0, tzinfo=timezone.utc)
    filter_timestamp = int(filter_date.timestamp())
    
    params = {
        'limit': 100,
        'created[gte]': filter_timestamp
    }
    
    processed_data = []
    
    with tqdm(desc="Fetching Stripe invoices") as pbar:
        while True:
            response = requests.get(stripe_base_url, headers=stripe_headers, params=params)
            
            if response.status_code != 200:
                print(f"❌ Failed to fetch Stripe data: Status code {response.status_code}")
                print(f"Response: {response.text}")
                break
            
            response_data = response.json()
            invoices = response_data.get('data', [])
            
            for invoice in invoices:
                lines_data = invoice['lines']['data']
                if lines_data:
                    first_line_item = lines_data[0]
                    lines_data_description = first_line_item.get('description', 'No description')
                    
                    period = first_line_item.get('period', {})
                    plan = first_line_item.get('plan', {})
                    
                    plan_active = plan.get('active', 'No plan active info') if plan else 'No plan active info'
                    plan_interval = plan.get('interval', 'No plan interval info') if plan else 'No plan interval info'
                else:
                    lines_data_description = 'No description'
                    plan_active = 'No plan active info'
                    plan_interval = 'No plan interval info'
                
                invoice_data = {
                    'id': invoice['id'],
                    'amount_due': invoice['amount_due'] / 100,
                    'amount_paid': invoice['amount_paid'] / 100,
                    'amount_remaining': invoice['amount_remaining'] / 100,
                    'created': convert_timestamp_to_date(invoice['created']),
                    'customer': invoice['customer'],
                    'lines_data_description': lines_data_description,
                    'plan_active': plan_active,
                    'plan_interval': plan_interval,
                    'subscription': invoice['subscription'],
                    'attempt_count': invoice.get('attempt_count', 0),
                    'payment_intent': invoice.get('payment_intent', 'No payment intent'),
                    'status': invoice['status'],
                    'paid': invoice.get('paid', False)
                }
                processed_data.append(invoice_data)
            
            pbar.update(len(invoices))
            
            if response_data.get('has_more', False) and invoices:
                last_invoice_id = invoices[-1]['id']
                params['starting_after'] = last_invoice_id
            else:
                break
    
    return processed_data

# =============================================================================
# MAIN DATA FETCHING
# =============================================================================

print("🚀 Starting enhanced data fetching process...")

# Fetch Calendesk data
print("\n" + "="*50)
print("FETCHING CALENDESK DATA")
print("="*50)

print("🔄 Fetching subscriptions data...")
all_subscriptions = fetch_calendesk_data(subscriptions_url, calendesk_headers, "Calendesk subscriptions")
validate_calendesk_data(all_subscriptions, 'subscriptions')

print("🔄 Fetching users subscriptions data...")
users_subscriptions = fetch_calendesk_data(users_url, calendesk_headers, "Calendesk users")
validate_calendesk_data(users_subscriptions, 'users')

print(f"✓ Calendesk data fetched: {len(all_subscriptions)} subscriptions, {len(users_subscriptions)} user subscriptions")

# Process Calendesk data
print("🔄 Processing Calendesk data...")

if not all_subscriptions:
    print("❌ No subscriptions data fetched - cannot continue")
    exit()

if not users_subscriptions:
    print("❌ No user subscriptions data fetched - cannot continue")
    exit()

try:
    subscriptions_df = pd.json_normalize(all_subscriptions, sep='.')
    users_subscriptions_df = pd.json_normalize(users_subscriptions, sep='.')
    
    print(f"✓ DataFrames created: {len(subscriptions_df)} subscriptions, {len(users_subscriptions_df)} user records")
    
    # Validate essential columns exist
    required_subscription_cols = ['id', 'price.recurring_interval']
    required_user_cols = ['id', 'subscription_id', 'user.id', 'status', 'stripe_subscription_id']
    
    missing_sub_cols = [col for col in required_subscription_cols if col not in subscriptions_df.columns]
    missing_user_cols = [col for col in required_user_cols if col not in users_subscriptions_df.columns]
    
    if missing_sub_cols:
        print(f"⚠ Warning: Missing subscription columns: {missing_sub_cols}")
        print(f"   Available columns: {list(subscriptions_df.columns)[:10]}...")
    
    if missing_user_cols:
        print(f"⚠ Warning: Missing user subscription columns: {missing_user_cols}")
        print(f"   Available columns: {list(users_subscriptions_df.columns)[:10]}...")
        
except Exception as e:
    print(f"❌ Error processing Calendesk data: {e}")
    print("   This might indicate a change in the API response structure")
    exit()

# Extract phone number
users_subscriptions_df['user_default_phone_e164'] = users_subscriptions_df['user.default_phone.e164'].fillna('')

# Merge DataFrames
print("🔄 Merging Calendesk DataFrames...")
df_calendesk = pd.merge(
    users_subscriptions_df,
    subscriptions_df[['id', 'price.recurring_interval']],
    left_on='subscription_id',
    right_on='id',
    how='left'
)

# Filter data
excluded_subscription_ids = [260, 231, 169, 157, 140, 92, 42, 9, 7]
excluded_user_ids = [9771, 9799, 10735, 9817, 100, 12113, 10860, 12216, 7185, 12218, 8819, 10635, 7921, 14480, 15416]

df_calendesk = df_calendesk[
    ~df_calendesk['id_y'].isin(excluded_subscription_ids) &
    df_calendesk['status'].isin(['active', 'canceled']) &
    ~df_calendesk['user.id'].isin(excluded_user_ids)
]

print(f"✓ Filtered and transformed Calendesk data: {len(df_calendesk)} rows")

# Transform Calendesk data
df_calendesk['Imię i nazwisko'] = df_calendesk['user.name'] + ' ' + df_calendesk['user.surname']
df_calendesk['default_address_name'] = df_calendesk['user.default_address.name'].fillna('')
df_calendesk['status'] = df_calendesk['status'].replace({'canceled': 'anulowana', 'active': 'aktywna'})
df_calendesk['price.recurring_interval'] = df_calendesk['price.recurring_interval'].replace({'year': 'roczny', 'month': 'miesięczny'})

# Select and rename columns for Calendesk
calendesk_columns = {
    'id_x': 'ID Subskrypcji Klienta',
    'id_y': 'ID Subskrypcji',
    'status': 'Status',
    'created_at': 'Data zakupu',
    'subscription.name': 'Pakiet',
    'ends_at': 'Data wygaśnięcia',
    'canceled_at': 'Data anulowania',
    'user.id': 'ID Klienta',
    'Imię i nazwisko': 'Imię i Nazwisko Klienta',
    'user.email': 'Email',
    'price.recurring_interval': 'Typ pakietu',
    'stripe_subscription_id': 'ID Suba STRIPE',
    'user.default_address.tax_number': 'NIP',
    'default_address_name': 'Nazwa Firmy',
    'user_default_phone_e164': 'Telefon'
}

df_calendesk = df_calendesk[list(calendesk_columns.keys())].rename(columns=calendesk_columns)

# Convert dates
df_calendesk['Data zakupu'] = pd.to_datetime(df_calendesk['Data zakupu']).dt.tz_localize(None) + pd.DateOffset(hours=2)
df_calendesk['Data anulowania'] = pd.to_datetime(df_calendesk['Data anulowania'], errors='coerce').dt.tz_localize(None) + pd.DateOffset(hours=2)
df_calendesk['Data wygaśnięcia'] = pd.to_datetime(df_calendesk['Data wygaśnięcia'], errors='coerce').dt.tz_localize(None) + pd.DateOffset(hours=2)

# Update cancellation dates
mask = df_calendesk['Data anulowania'].isna() & df_calendesk['Data wygaśnięcia'].notna()
df_calendesk.loc[mask, 'Data anulowania'] = df_calendesk['Data wygaśnięcia']

# Process NIP column - convert to numeric if possible, keep as text if it contains formatting
def process_nip(nip_value):
    """Convert NIP to number if it's purely numeric, otherwise keep as text"""
    if pd.isna(nip_value) or nip_value == '':
        return nip_value
    
    nip_str = str(nip_value).strip()
    
    # If it contains only digits, convert to number
    if nip_str.isdigit():
        return int(nip_str)
    
    # If it contains formatting characters (like dashes), keep as text
    return nip_str

df_calendesk['NIP'] = df_calendesk['NIP'].apply(process_nip)

print("✓ Calendesk data processing completed")

# Fetch Stripe data
print("\n" + "="*50)
print("FETCHING STRIPE DATA")
print("="*50)

stripe_data = fetch_stripe_invoices()
df_stripe = pd.DataFrame(stripe_data)

if not df_stripe.empty:
    df_stripe['created'] = pd.to_datetime(df_stripe['created'])
    print(f"✓ Stripe data fetched and processed: {len(df_stripe)} invoices")
    
    # Rename columns for Stripe data
    stripe_columns = {
        'id': 'ID_Faktury',
        'amount_due': 'Kwota Do Zapłaty',
        'amount_paid': 'Kwota Zapłacona',
        'amount_remaining': 'Pozostało Do Zapłaty',
        'created': 'Data Utworzenia',
        'customer': 'ID_Klienta',
        'lines_data_description': 'Pakiet',
        'plan_active': 'Sub Aktywny',
        'plan_interval': 'Okres Odnowienia',
        'subscription': 'ID_Subskrypcji',
        'attempt_count': 'Liczba Pobrań Kwoty Do Zapłaty',
        'payment_intent': 'ID_Payment_Intent',
        'status': 'Status Faktury',
        'paid': 'Faktura Opłacona'
    }
    
    df_stripe = df_stripe.rename(columns=stripe_columns)
    print(f"✓ Stripe columns mapped: {len(df_stripe)} invoices ready")
else:
    print("⚠ No Stripe data found")

# =============================================================================
# CALCULATE CUSTOM COLUMNS FOR CALENDESK DATA
# =============================================================================

print("\n🔄 Calculating custom columns...")

def calculate_invoice_status_chosen_month(row, invoices_df, chosen_month=6, chosen_year=2025):
    """Calculate invoice status for chosen month"""
    if pd.isna(row['ID Suba STRIPE']) or row['ID Suba STRIPE'] == '':
        return "Nie można określić"
    
    subscription_id = row['ID Suba STRIPE']
    package_type = row['Typ pakietu']
    
    if package_type == 'miesięczny':
        # Monthly subscription
        filtered_invoices = invoices_df[
            (invoices_df['ID_Subskrypcji'] == subscription_id) &
            (invoices_df['Data Utworzenia'].dt.month == chosen_month) &
            (invoices_df['Data Utworzenia'].dt.year == chosen_year)
        ]
        if not filtered_invoices.empty:
            return filtered_invoices.iloc[0]['Status Faktury']
    elif package_type == 'roczny':
        # Yearly subscription
        start_date = dt(2024, chosen_month, 1)
        end_date = dt(chosen_year, chosen_month, 30)
        filtered_invoices = invoices_df[
            (invoices_df['ID_Subskrypcji'] == subscription_id) &
            (invoices_df['Data Utworzenia'] >= start_date) &
            (invoices_df['Data Utworzenia'] <= end_date)
        ]
        if not filtered_invoices.empty:
            return filtered_invoices.iloc[0]['Status Faktury']
    
    return ""

def calculate_invoice_status_last_2_months(row, invoices_df, month1=6, month2=7, year=2025):
    """Calculate invoice status for last 2 months"""
    if pd.isna(row['ID Suba STRIPE']) or row['ID Suba STRIPE'] == '':
        return "Nie można określić"
    
    subscription_id = row['ID Suba STRIPE']
    package_type = row['Typ pakietu']
    
    if package_type == 'miesięczny':
        # Monthly subscription
        start_date = dt(year, month1, 1)
        end_date = dt(year, month2, 31)
        filtered_invoices = invoices_df[
            (invoices_df['ID_Subskrypcji'] == subscription_id) &
            (invoices_df['Data Utworzenia'] >= start_date) &
            (invoices_df['Data Utworzenia'] <= end_date) &
            (invoices_df['Status Faktury'] == 'paid')
        ]
        return "paid" if not filtered_invoices.empty else ""
    elif package_type == 'roczny':
        # Yearly subscription
        start_date = dt(2024, month2, 1)
        end_date = dt(year, month2, 31)
        filtered_invoices = invoices_df[
            (invoices_df['ID_Subskrypcji'] == subscription_id) &
            (invoices_df['Data Utworzenia'] >= start_date) &
            (invoices_df['Data Utworzenia'] <= end_date) &
            (invoices_df['Status Faktury'] == 'paid')
        ]
        return "paid" if not filtered_invoices.empty else ""
    
    return ""

def calculate_last_invoice_month(row, invoices_df, current_year=2025):
    """Calculate last invoice month"""
    if pd.isna(row['ID Suba STRIPE']) or row['ID Suba STRIPE'] == '':
        return ""
    
    subscription_id = row['ID Suba STRIPE']
    package_type = row['Typ pakietu']
    
    filtered_invoices = invoices_df[
        (invoices_df['ID_Subskrypcji'] == subscription_id) &
        (invoices_df['Status Faktury'] == 'paid')
    ]
    
    if package_type == 'miesięczny':
        filtered_invoices = filtered_invoices[
            filtered_invoices['Data Utworzenia'].dt.year == current_year
        ]
    elif package_type == 'roczny':
        filtered_invoices = filtered_invoices[
            filtered_invoices['Data Utworzenia'] >= dt(2024, 1, 1)
        ]
    
    if not filtered_invoices.empty:
        max_date = filtered_invoices['Data Utworzenia'].max()
        return max_date.strftime('%B').lower()  # Return month name in Polish would need translation
    
    return ""

# Apply custom calculations if we have Stripe data
if not df_stripe.empty:
    print("  ✓ Calculating invoice status for chosen month...")
    df_calendesk['Invoice status in chosen month'] = df_calendesk.apply(
        lambda row: calculate_invoice_status_chosen_month(row, df_stripe), axis=1
    )
    
    print("  ✓ Calculating invoice status for last 2 months...")
    df_calendesk['Invoice status in last 2 months'] = df_calendesk.apply(
        lambda row: calculate_invoice_status_last_2_months(row, df_stripe), axis=1
    )
    
    print("  ✓ Calculating last invoice month...")
    df_calendesk['Last invoice month'] = df_calendesk.apply(
        lambda row: calculate_last_invoice_month(row, df_stripe), axis=1
    )
else:
    # Add empty columns if no Stripe data
    df_calendesk['Invoice status in chosen month'] = ""
    df_calendesk['Invoice status in last 2 months'] = ""
    df_calendesk['Last invoice month'] = ""

print("✓ Custom columns calculated")

# =============================================================================
# SAVE TO EXCEL
# =============================================================================

print("\n" + "="*50)
print("SAVING TO EXCEL")
print("="*50)

print(f"📂 Loading Excel workbook: {excel_file_path}")

try:
    if not os.path.exists(excel_file_path):
        print(f"❌ File does not exist: {excel_file_path}")
        exit()
    
    wb = load_workbook(excel_file_path)
    print("✓ Workbook loaded successfully")
    
    # Update Calendesk sheet
    if calendesk_sheet_name in wb.sheetnames:
        print(f"✓ Found sheet: {calendesk_sheet_name}")
        ws_calendesk = wb[calendesk_sheet_name]
        
        # Clear existing data (starting from row 2 to preserve headers)
        print("🧹 Clearing existing Calendesk data...")
        max_row = ws_calendesk.max_row
        max_col = len(df_calendesk.columns)
        
        for row in ws_calendesk.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.value = None
        
        # Write new data
        print(f"✍ Writing Calendesk data ({len(df_calendesk)} rows)...")
        data_to_write = df_calendesk.values.tolist()
        
        for r_idx, row in enumerate(data_to_write, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_calendesk.cell(row=r_idx, column=c_idx, value=value)
        
        print(f"✓ Calendesk data written to {calendesk_sheet_name}")
    else:
        print(f"❌ Sheet {calendesk_sheet_name} not found")
    
    # Update Stripe sheet
    if not df_stripe.empty and stripe_sheet_name in wb.sheetnames:
        print(f"✓ Found sheet: {stripe_sheet_name}")
        ws_stripe = wb[stripe_sheet_name]
        
        # Clear existing data
        print("🧹 Clearing existing Stripe data...")
        max_row = ws_stripe.max_row
        max_col = len(df_stripe.columns)
        
        for row in ws_stripe.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.value = None
        
        # Write new data
        print(f"✍ Writing Stripe data ({len(df_stripe)} rows)...")
        data_to_write = df_stripe.values.tolist()
        
        for r_idx, row in enumerate(data_to_write, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_stripe.cell(row=r_idx, column=c_idx, value=value)
        
        print(f"✓ Stripe data written to {stripe_sheet_name}")
    elif df_stripe.empty:
        print("⚠ No Stripe data to write")
    else:
        print(f"❌ Sheet {stripe_sheet_name} not found")
    
    # Save workbook
    print("💾 Saving workbook...")
    wb.save(excel_file_path)
    print("✅ Workbook saved successfully")

except Exception as e:
    print(f"❌ Error working with Excel file: {e}")
    exit()

# =============================================================================
# FINAL SUMMARY
# =============================================================================

print("\n" + "="*50)
print("🎉 PROCESS COMPLETED SUCCESSFULLY!")
print("="*50)
print(f"📊 Final summary:")
print(f"  - Calendesk records processed: {len(df_calendesk)}")
print(f"  - Stripe records processed: {len(df_stripe) if not df_stripe.empty else 0}")
print(f"  - File updated: {excel_file_path}")
print(f"  - Sheets updated: {calendesk_sheet_name}" + (f", {stripe_sheet_name}" if not df_stripe.empty else ""))
print("="*50) 