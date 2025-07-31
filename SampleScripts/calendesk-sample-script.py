import os
import requests
import pandas as pd
import datetime
from tqdm import tqdm
from openpyxl import load_workbook

# Set the API key and endpoint for bookings
api_key = os.getenv('CALENDESK_API_KEY')
if not api_key:
    print("No API key found in environment. Please set the CALENDESK_API_KEY variable.")
    exit()

print(f"✓ API key loaded successfully")

# Endpoints and headers
subscriptions_url = 'https://api.calendesk.com/api/admin/subscriptions'
users_url = 'https://api.calendesk.com/api/admin/v2/users/subscriptions'
headers = {
    "X-Tenant": "slawomir-mentzen-rvs",
    "X-Api-Key": api_key
}

print(f"✓ API endpoints configured")
print(f"  - Subscriptions URL: {subscriptions_url}")
print(f"  - Users URL: {users_url}")

# Function to fetch data from a specified number of pages
def fetch_pages_data(url, headers, pages_to_fetch):
    all_data = []
    print(f"📡 Starting to fetch data from {url} ({pages_to_fetch} pages)")
    with tqdm(desc="Pobieranie danych", unit="strona") as pbar:
        for page in range(1, pages_to_fetch + 1):
            params = {
                'limit': 100,
                'page': page,
                'order_by': 'id',
                'ascending': 0
            }
            response = requests.get(url, headers=headers, params=params)
            if response.status_code == 200:
                data = response.json().get('data', [])
                if data:
                    all_data.extend(data)
                    pbar.update(1)
                    print(f"  ✓ Page {page}: {len(data)} records fetched")
                else:
                    print(f"  ⚠ Page {page}: No data returned, stopping")
                    break
            else:
                print(f'  ❌ Błąd zapytania na stronie {page}: {response.status_code}')
                break
    print(f"✓ Total records fetched: {len(all_data)}")
    return all_data

# Fetch data from 10 pages for all_subscriptions and 1000 pages for users_subscriptions
print("\n🔄 Fetching subscriptions data...")
all_subscriptions = fetch_pages_data(subscriptions_url, headers, 10)

print("\n🔄 Fetching users subscriptions data...")
users_subscriptions = fetch_pages_data(users_url, headers, 1000)

# Transform fetched data into DataFrame
print("\n🔄 Processing data...")
subscriptions_df = pd.json_normalize(all_subscriptions, sep='.')
users_subscriptions_df = pd.json_normalize(users_subscriptions, sep='.')

print(f"✓ Subscriptions DataFrame created: {len(subscriptions_df)} rows")
print(f"✓ Users subscriptions DataFrame created: {len(users_subscriptions_df)} rows")

# Extract the e164 field for user's default phone
users_subscriptions_df['user_default_phone_e164'] = users_subscriptions_df['user.default_phone.e164'].fillna('')

# Display the number of subscriptions fetched
print(f"Łączna liczba subskrypcji pobranych z endpointu subscriptions: {len(all_subscriptions)}")
print(f"Łączna liczba subskrypcji pobranych z endpointu users/subscriptions: {len(users_subscriptions)}")

# Merge DataFrames
print("🔄 Merging DataFrames...")
df = pd.merge(
    users_subscriptions_df,
    subscriptions_df[['id', 'price.recurring_interval']],
    left_on='subscription_id',
    right_on='id',
    how='left'
)
print(f"✓ DataFrames merged: {len(df)} rows")

# Exclude specific subscription IDs
excluded_subscription_ids = [260, 231, 169, 157, 140, 92, 42, 9, 7]
df_before_exclusion = len(df)
df = df[~df['id_y'].isin(excluded_subscription_ids)]
print(f"🔄 Excluded subscription IDs: {excluded_subscription_ids}")
print(f"✓ Rows after subscription ID exclusion: {len(df)} (removed: {df_before_exclusion - len(df)})")

# Filter records to exclude specific user IDs
excluded_user_ids = [9771, 9799, 10735, 9817, 100, 12113, 10860, 12216, 7185, 12218, 8819, 10635, 7921, 14480, 15416]
df_before_user_exclusion = len(df)
df = df[df['status'].isin(['active', 'canceled']) & ~df['user.id'].isin(excluded_user_ids)]
print(f"🔄 Excluded user IDs: {excluded_user_ids}")
print(f"✓ Rows after user ID exclusion and status filter: {len(df)} (removed: {df_before_user_exclusion - len(df)})")

# Display the number of subscriptions after filtering
print(f"Łączna liczba subskrypcji po przefiltrowaniu: {len(df)}")

# Combine first and last names into a single column
df['Imię i nazwisko'] = df['user.name'] + ' ' + df['user.surname']

# Extract the default_address_name field
df['default_address_name'] = df['user.default_address.name'].fillna('')

# Change status and package type names
df['status'] = df['status'].replace({'canceled': 'anulowana', 'active': 'aktywna'})
df['price.recurring_interval'] = df['price.recurring_interval'].replace({'year': 'roczny', 'month': 'miesięczny'})

print("✓ Data transformation completed")

# Define columns to keep and rename
kolumny_do_zachowania = [
    'id_x', 'id_y', 'status', 'created_at', 'subscription.name',
    'ends_at', 'canceled_at', 'user.id', 'Imię i nazwisko',
    'user.email', 'price.recurring_interval', 'stripe_subscription_id',
    'user.default_address.tax_number', 'default_address_name',
    'user_default_phone_e164'
]

df = df[kolumny_do_zachowania].rename(columns={
    'id_x': 'ID Subskrypcji Klienta',
    'id_y': 'ID Subskrypcji',
    'status': 'Status',
    'created_at': 'Data zakupu',
    'subscription.name': 'Pakiet',
    'ends_at': 'Data wygaśnięcia',
    'canceled_at': 'Data anulowania',
    'user.id': 'ID Klienta',
    'Imię i nazwisko': 'Imię i nazwisko',
    'user.email': 'Email',
    'price.recurring_interval': 'Typ pakietu',
    'stripe_subscription_id': 'ID Suba STRIPE',
    'user.default_address.tax_number': 'NIP',
    'default_address_name': 'Nazwa adresu',
    'user_default_phone_e164': 'Telefon'
})

print(f"✓ Columns filtered and renamed: {len(df.columns)} columns, {len(df)} rows")

# Convert dates to datetime objects and remove timezone information
print("🔄 Converting dates...")
df['Data zakupu'] = pd.to_datetime(df['Data zakupu']).dt.tz_localize(None) + pd.DateOffset(hours=2)
df['Data anulowania'] = pd.to_datetime(df['Data anulowania'], errors='coerce').dt.tz_localize(None) + pd.DateOffset(hours=2)
df['Data wygaśnięcia'] = pd.to_datetime(df['Data wygaśnięcia'], errors='coerce').dt.tz_localize(None) + pd.DateOffset(hours=2)

def update_cancellation_dates(df):
    mask = df['Data anulowania'].isna() & df['Data wygaśnięcia'].notna()
    df.loc[mask, 'Data anulowania'] = df['Data wygaśnięcia']
    return df, mask

df, _ = update_cancellation_dates(df)
print("✓ Date conversion and cancellation date update completed")

# Convert DataFrame to list of lists for inserting into Excel
data_to_insert = df.values.tolist()
print(f"✓ Data prepared for Excel: {len(data_to_insert)} rows")

# Loading the workbook and selecting the sheet
save_path = r'C:\Users\w.kuczkowski\OneDrive - Sławomir Mentzen\Pulpit\Baza subskrypcji.xlsx'
sheet_name = 'Dane'

print(f"\n📂 Loading Excel workbook...")
print(f"  File path: {save_path}")
print(f"  Sheet name: {sheet_name}")

try:
    # Check if file exists
    if os.path.exists(save_path):
        print(f"✓ File exists")
    else:
        print(f"❌ File does not exist at path: {save_path}")
        exit()

    wb = load_workbook(save_path)
    print(f"✓ Workbook loaded successfully")
    
    # Check if sheet exists
    if sheet_name in wb.sheetnames:
        print(f"✓ Sheet '{sheet_name}' found")
        ws = wb[sheet_name]
    else:
        print(f"❌ Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
        exit()

    print(f"✓ Current sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")

except Exception as e:
    print(f"❌ Error loading workbook: {e}")
    exit()

# Clear existing data only from columns A to O
print("\n🧹 Clearing existing data from columns A to O...")
cleared_cells = 0
for row in ws.iter_rows(min_row=2, min_col=1, max_col=15, max_row=ws.max_row):
    for cell in row:
        if cell.value is not None:
            cleared_cells += 1
        cell.value = None

print(f"✓ Cleared {cleared_cells} cells")

# Write new data
print(f"\n✍ Writing new data to Excel...")
print(f"  Data to write: {len(data_to_insert)} rows x {len(data_to_insert[0]) if data_to_insert else 0} columns")

written_cells = 0
for r_idx, row in enumerate(data_to_insert, start=2):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)
        written_cells += 1
    if r_idx % 100 == 0:  # Progress update every 100 rows
        print(f"  ✓ Written {r_idx - 1} rows...")

print(f"✓ Written {written_cells} cells total")

# Save the workbook
print(f"\n💾 Saving workbook...")
try:
    wb.save(save_path)
    print(f"✅ Workbook saved successfully to: {save_path}")
except Exception as e:
    print(f"❌ Error saving workbook: {e}")
    exit()

print(f"\n🎉 Process completed successfully!")
print(f"📊 Final summary:")
print(f"  - Records processed: {len(df)}")
print(f"  - Rows written to Excel: {len(data_to_insert)}")
print(f"  - Columns written: {len(df.columns)}")
print(f"  - File saved to: {save_path}")