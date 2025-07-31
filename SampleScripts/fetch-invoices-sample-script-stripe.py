import os
import requests
import pandas as pd
from datetime import datetime, timezone
from tqdm import tqdm
from openpyxl import load_workbook

# Set your secret API key here
stripe_api_key = os.environ.get('STRIPE_API_KEY')

# The base URL for the Stripe API
base_url = 'https://api.stripe.com/v1/invoices'

# Headers for authorization
headers = {
    'Authorization': f'Bearer {stripe_api_key}',
}

# Calculate Unix timestamp for June 1st, 2025
filter_date = datetime(2025, 6, 1, 0, 0, 0, tzinfo=timezone.utc)
filter_timestamp = int(filter_date.timestamp())

# Initialize parameters for pagination with date filter
params = {
    'limit': 100,  # Number of results per page.
    'created[gte]': filter_timestamp  # Only retrieve invoices created on or after June 1st, 2025
}

# Function to convert Unix timestamp to a date without timezone
def convert_timestamp_to_date(timestamp):
    return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d')

# Initialize a list to store the processed data
processed_data = []
# Initialize a tqdm progress bar
with tqdm(desc="Fetching invoices from 01.06.2025") as pbar:
    # Make API calls to fetch all pages of data
    while True:
        response = requests.get(base_url, headers=headers, params=params)
        # Check if the response was successful
        if response.status_code != 200:
            print(f"Failed to fetch data: Status code {response.status_code}")
            print(f"Response: {response.text}")
            break
        
        response_data = response.json()
        invoices = response_data.get('data', [])

        # Process each invoice
        for invoice in invoices:
            lines_data = invoice['lines']['data']
            if lines_data:  # Check if 'lines' 'data' is not empty
                first_line_item = lines_data[0]
                lines_data_description = first_line_item.get('description', 'No description')
                
                # Check if 'period' and 'plan' exist and are not None before accessing sub-keys
                period = first_line_item.get('period', {})
                plan = first_line_item.get('plan', {})
                
                plan_active = plan.get('active', 'No plan active info') if plan else 'No plan active info'
                plan_interval = plan.get('interval', 'No plan interval info') if plan else 'No plan interval info'
            else:  # If 'lines' 'data' is empty, set defaults or skip
                lines_data_description = 'No description'
                plan_active = 'No plan active info'
                plan_interval = 'No plan interval info'

            # Prepare the invoice data with all the gathered information and the new fields
            invoice_data = {
                'id': invoice['id'],
                'amount_due': invoice['amount_due'] / 100,
                'amount_paid': invoice['amount_paid'] / 100,
                'amount_remaining': invoice['amount_remaining'] / 100,
                'created': convert_timestamp_to_date(invoice['created']),
                'customer': invoice['customer'],
                'lines_data_description': lines_data_description,
                'plan_active': plan_active,
                'plan_interval': plan_interval,
                'subscription': invoice['subscription'],
                'attempt_count': invoice.get('attempt_count', 0),  # Default to 0 if not present
                'payment_intent': invoice.get('payment_intent', 'No payment intent'),  # Default to a placeholder
                'status': invoice['status'],
                'paid': invoice.get('paid', False)  # Default to False if not present
            }
            processed_data.append(invoice_data)
        pbar.update(len(invoices))  # Update progress bar with number of invoices fetched

        # Prepare for the next page
        if response_data.get('has_more', False) and invoices:
            last_invoice_id = invoices[-1]['id']
            params['starting_after'] = last_invoice_id
        else:
            break

# Convert the processed data to a pandas DataFrame
df = pd.DataFrame(processed_data)

# Convert date columns to datetime format
df['created'] = pd.to_datetime(df['created'])

# Define the file path and sheet name
save_path = r'C:\Users\w.kuczkowski\OneDrive - Sławomir Mentzen\Pulpit\Baza subskrypcji.xlsx'
sheet_name = 'Dane (Faktury_Stripe)'

# Load the existing workbook
wb = load_workbook(save_path)
ws = wb[sheet_name]

# Clear existing data
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# Write the new data starting from A2
for r_idx, row in enumerate(df.itertuples(index=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)

# Save the workbook
wb.save(save_path) 